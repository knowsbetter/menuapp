from openpyxl import Workbook


def convert_to_excel(res: list):
    wb = Workbook()
    ws = wb.active

    menu_rows = []
    submenu_rows = []

    for i, r in enumerate(res[0]):
        row = i * (r["dishes_count"] + 3) + 1
        ws.cell(row=row, column=1, value=i + 1)
        ws.cell(row=row, column=2, value=r["title"])
        ws.cell(row=row, column=3, value=r["description"])
        menu_rows.append(row)

    d = 0
    num = 0

    for i, r in enumerate(res[1]):
        row = i * (r["dishes_count"] + 1) + 2 + d
        if row in menu_rows:
            row += 1
            num = 0
        ws.cell(row=row, column=2, value=num + 1)
        ws.cell(row=row, column=3, value=r["title"])
        ws.cell(row=row, column=4, value=r["description"])
        d = row - i * (r["dishes_count"] + 1) - 2
        num += 1
        submenu_rows.append(row)

    d = 0
    num = 0

    for i, r in enumerate(res[2]):
        row = i + 3 + d
        row = row + 1 if row in menu_rows else row
        if row in submenu_rows:
            row = row + 1
            num = 0
        ws.cell(row=row, column=3, value=num + 1)
        ws.cell(row=row, column=4, value=r["title"])
        ws.cell(row=row, column=5, value=r["description"])
        ws.cell(row=row, column=6, value=r["price"])
        num += 1
        d = row - i - 3

    wb.save("output/test_menu.xlsx")
